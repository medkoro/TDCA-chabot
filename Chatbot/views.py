import os
import json
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import pandas as pd
import logging
from datetime import datetime, date
import numpy as np
import matplotlib.pyplot as plt
import io, base64
from langchain.prompts import PromptTemplate
from langchain_google_genai import ChatGoogleGenerativeAI
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def index(request):
    # Check if there's any uploaded file in session
    uploaded_file = request.session.get('uploaded_file', None)
    sheet_names = request.session.get('sheet_names', [])
    current_sheet = request.session.get('current_sheet', '')
    
    context = {
        'uploaded_file': uploaded_file,
        'sheet_names': sheet_names,
        'current_sheet': current_sheet,
    }
    return render(request, 'chatbot.html', context)

@csrf_exempt
def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Save the file temporarily
        file_path = os.path.join(settings.MEDIA_ROOT, 'temp', excel_file.name)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'wb+') as destination:
            for chunk in excel_file.chunks():
                destination.write(chunk)
        
        # Get sheet names
        if excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xlsm'):
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        else:  # For .xls files
            xls = pd.ExcelFile(file_path, engine='xlrd')
            sheet_names = xls.sheet_names
            xls.close()
        
        # Store file info in session
        request.session['uploaded_file'] = excel_file.name
        request.session['file_path'] = file_path
        request.session['sheet_names'] = sheet_names
        request.session['current_sheet'] = sheet_names[0] if sheet_names else ''
        
        return JsonResponse({
            'status': 'success',
            'file_name': excel_file.name,
            'sheet_names': sheet_names,
            'current_sheet': sheet_names[0] if sheet_names else ''
        })
    
    return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)

def get_sheet_data(request, sheet_name):
    if 'file_path' not in request.session:
        return JsonResponse({'status': 'error', 'message': 'No file loaded'}, status=400)
    
    try:
        file_path = request.session['file_path']
        
        # Read the Excel file
        if file_path.endswith(('.xlsx', '.xlsm')):
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        else:  # For .xls files
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
        
        # Convert DataFrame to JSON
        data = {
            'columns': df.columns.tolist(),
            'data': df.where(pd.notnull(df), None).values.tolist(),
            'current_sheet': sheet_name
        }
        
        # Update current sheet in session
        request.session['current_sheet'] = sheet_name
        
        return JsonResponse({'status': 'success', 'data': data})
    
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
def generate_graph(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('fileInput')
        command = request.POST.get('command')
        
        # Check if we have session data instead of uploaded file
        if not excel_file and 'file_path' in request.session:
            file_path = request.session['file_path']
            sheet_name = request.session.get('current_sheet', '')
            try:
                if file_path.endswith(('.xlsx', '.xlsm')):
                    xls = pd.ExcelFile(file_path, engine='openpyxl')
                    if not sheet_name:
                        sheet_name = xls.sheet_names[0]
                    df = pd.read_excel(xls, sheet_name=sheet_name, nrows=5000)
                elif file_path.endswith('.xls'):
                    xls = pd.ExcelFile(file_path, engine='xlrd')
                    if not sheet_name:
                        sheet_name = xls.sheet_names[0]
                    df = pd.read_excel(xls, sheet_name=sheet_name, nrows=5000)
                else:
                    df = pd.read_csv(file_path)
                    sheet_name = 'csv'
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Erreur de lecture du fichier: {str(e)}'})
        elif excel_file:
            if hasattr(excel_file, 'seek'):
                excel_file.seek(0)
            file_name = excel_file.name.lower()

            try:
                if file_name.endswith('.csv'):
                    try:
                        df = pd.read_csv(excel_file)
                    except UnicodeDecodeError:
                        excel_file.seek(0)
                        df = pd.read_csv(excel_file, encoding='latin1')
                    sheet_name = 'csv'
                elif file_name.endswith(('.xlsx', '.xlsm')):
                    xls = pd.ExcelFile(excel_file, engine='openpyxl')
                    sheet_name = xls.sheet_names[0]
                    for name in xls.sheet_names:
                        if name.lower() in command.lower():
                            sheet_name = name
                            break
                    df = pd.read_excel(xls, sheet_name=sheet_name, nrows=5000)
                elif file_name.endswith('.xls'):
                    xls = pd.ExcelFile(excel_file, engine='xlrd')
                    sheet_name = xls.sheet_names[0]
                    for name in xls.sheet_names:
                        if name.lower() in command.lower():
                            sheet_name = name
                            break
                    df = pd.read_excel(xls, sheet_name=sheet_name, nrows=5000)
                else:
                    return JsonResponse({'status': 'error', 'message': 'Format de fichier non supporté. Veuillez fournir un fichier .csv, .xlsx ou .xls.'})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Erreur de lecture du fichier: {str(e)}'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Aucun fichier disponible.'})

        # Analyze data to suggest appropriate visualizations
        numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
        categorical_columns = df.select_dtypes(include=['object']).columns.tolist()
        
        cmd = command.lower()
        
        # Smart chart selection based on data analysis
        if "barre" in cmd or (not cmd and len(numeric_columns) > 0 and len(categorical_columns) > 0):
            ax = df.plot(kind='bar')
        elif "courbe" in cmd or "line" in cmd or (not cmd and len(numeric_columns) > 1):
            ax = df.plot(kind='line')
        elif "secteur" in cmd or "pie" in cmd or (not cmd and len(categorical_columns) > 0 and len(df[categorical_columns[0]].unique()) < 10):
            if categorical_columns:
                df[categorical_columns[0]].value_counts().plot.pie(autopct='%1.1f%%')
            else:
                df.iloc[:, 0].value_counts().plot.pie(autopct='%1.1f%%')
        elif "histogramme" in cmd or (not cmd and len(numeric_columns) > 0):
            ax = df.plot(kind='hist')
        else:
            # Default: try to create the most appropriate chart
            if len(numeric_columns) > 1:
                ax = df.plot(kind='line')
            elif len(categorical_columns) > 0 and len(df[categorical_columns[0]].unique()) < 15:
                df[categorical_columns[0]].value_counts().plot.pie(autopct='%1.1f%%')
            else:
                ax = df.plot(kind='bar')

        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        plt.close()
        buf.seek(0)
        image_base64 = base64.b64encode(buf.read()).decode('utf-8')

        return JsonResponse({'status': 'success', 'image': image_base64, 'sheet': sheet_name})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'})

@csrf_exempt
def analyze_command(request):
    if request.method == "POST":
        text = request.POST.get("text", "")

        try:
            llm = ChatGoogleGenerativeAI(
                model="gemini-1.5-flash",
                google_api_key="AIzaSyDrefO2vmE1AFx1cvRRDMK9_nKxedu3EXE"
            )

            # Enhanced prompt for dashboard creation
            prompt = (
                "Tu es un expert en création de dashboards. Réponds à la question suivante en te concentrant sur la création de visualisations et d'analyses de données. "
                "Si on te demande de créer un dashboard, analyse les données disponibles et propose des visualisations pertinentes. "
                "Sois concis mais informatif.\n\n"
                f"Question : {text}"
            )
            result = llm.invoke(prompt)

            content = getattr(result, 'content', None)
            if content:
                return JsonResponse({"result": content})
            return JsonResponse({"result": str(result)})
        except Exception as e:
            return JsonResponse({"error": f"Erreur d'analyse: {str(e)}"}, status=500)

    return JsonResponse({"error": "Méthode non autorisée"}, status=405)
